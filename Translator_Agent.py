from phi.agent import Agent
from phi.tools.duckduckgo import DuckDuckGo
from phi.model.groq import Groq
from dotenv import load_dotenv
from phi.agent import Agent, RunResponse
import openpyxl as xl


file_name = "YOUR FILE NAME.xlsx"  # The name of your Excel file

fhandle = xl.load_workbook(f"C:\\Users\McKinley Alex\Desktop\My programs\Tools\Translator\\{file_name}")
sheet = fhandle.active

load_dotenv()

agent = Agent(
    model=Groq(id="llama-3.3-70b-versatile"),
    tools = [DuckDuckGo()],
    name = "Japanese to English Translator Agent",
    role = "Translate blocks of Japanese text into English",
    instructions = ["Search Google to confirm place names"],
)

def translate_chunk(japanese_text):
    response: RunResponse = agent.run(f"Your job is to translate chunks of Japanese text and return the English translations. Here is some text for you to translate: {japanese_text}. Please only return the translated text.")
    translation = response.content
    return translation


workbook = xl.load_workbook(file_name)
active_sheet = workbook.active

for row_index in range(1, active_sheet.max_row + 1):
    cell_value = active_sheet.cell(row=row_index, column=1).value

    if cell_value:
        text_to_translate = cell_value
        translated_text = translate_chunk(text_to_translate)
        active_sheet.cell(row=row_index, column=2, value=translated_text)
    else:
        continue


workbook.save(f"Translated {file_name}")
workbook.close()